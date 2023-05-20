import openpyxl


try:
    # Cria os templates dos arquivos de saída
    with open("C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\Output\\Não Encontrados.txt", 'w+') as file1:
        pass
    with open("C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\Output\\Negativos e Estoque Maior.txt", 'w+') as file2:
        pass

    # Cria um workbook para armazenar os dados do arquivo de vendas
    output_workbook = openpyxl.Workbook()

    # Função para escrever os arquivos de checagem
    def NotFound(text):
        with open("C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\Output\\Não Encontrados.txt", 'a') as file1:
            file1.write(text)
            file1.write("\n")
            file1.close()

    def NegativeAndHighStock(text):
        with open("C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\Output\\Negativos e Estoque Maior.txt", 'a') as file2:
            file2.write(text)
            file2.write("\n")
            file2.close()

    # Abre os arquivos
    workbookBmais = openpyxl.load_workbook('C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\BANCAMAIS.xlsx')
    workbookMercos = openpyxl.load_workbook('C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\MERCOS.xlsx')

    # Abre as planilhas
    sheetBmais = workbookBmais.active
    sheetMercos = workbookMercos.active
    output_sheet = output_workbook.active

    #Número da fila
    row_number = 1

    # Cria os dicionários que armazenarão os dados de cada planilha
    dictBmais = {}
    dictMercos = {}

    # Preenche os dicionários com os dados de cada planilha
    for row in sheetBmais.iter_rows(values_only=True):
        dictBmais[row[0]] = row[1]

    for row in sheetMercos.iter_rows(values_only=True):
        dictMercos[row[0]] = row[1]

    # Checa se os produtos do Bmais estão no Mercos e se o estoque é maior
    for key in dictBmais:
        if int(dictBmais[key]) < 0:
            NegativeAndHighStock(f"Produto do Bmais com estoque negativo: Cód: {key} - Estoque: {dictBmais[key]}")
        else:
            if str(key) in dictMercos:
                if int(dictMercos[str(key)]) < 0:
                    NegativeAndHighStock(f"Produto do Mercos com estoque negativo: Cód: {key} - Estoque: {dictMercos[str(key)]}")
                elif int(dictMercos[str(key)]) - int(dictBmais[key]) > 0:
                    diff = int(dictMercos[str(key)]) - int(dictBmais[key])
                    output_sheet.cell(row=row_number, column=1, value=key)
                    output_sheet.cell(row=row_number, column=2, value=diff)
                    row_number += 1
                elif int(dictMercos[str(key)]) - int(dictBmais[key]) < 0:
                    NegativeAndHighStock(f"Produto do Bmais com estoque MAIOR: Cód: {key} - [BMAIS] Estoque: {dictBmais[key]} // [MERCOS] Estoque: {dictMercos[str(key)]}")
            elif ("0" + str(key)) in dictMercos:
                key0 = "0" + str(key)
                if int(dictMercos[key0]) < 0:
                    NegativeAndHighStock(f"Produto do Mercos com estoque negativo: Cód: {key0} - Estoque: {dictMercos[key0]}")
                elif int(dictMercos[key0]) - int(dictBmais[key]) > 0:
                    diff = int(dictMercos[key0]) - int(dictBmais[key])
                    output_sheet.cell(row=row_number, column=1, value=key0)
                    output_sheet.cell(row=row_number, column=2, value=diff)
                    row_number += 1
                elif int(dictMercos[key0]) - int(dictBmais[key]) < 0:
                    NegativeAndHighStock(f"Produto do Bmais com estoque MAIOR: Cód: {key0} - [BMAIS] Estoque: {dictBmais[key]} // [MERCOS] Estoque: {dictMercos[key0]}") 
            elif ("00" + str(key)) in dictMercos:
                key00 = "00" + str(key)
                if int(dictMercos[key00]) < 0:
                    NegativeAndHighStock(f"Produto do Mercos com estoque negativo: Cód: {key00} - Estoque: {dictMercos[key00]}")
                elif int(dictMercos[key00]) - int(dictBmais[key]) > 0:
                    diff = int(dictMercos[key00]) - int(dictBmais[key])
                    output_sheet.cell(row=row_number, column=1, value=key00)
                    output_sheet.cell(row=row_number, column=2, value=diff)
                    row_number += 1
                elif int(dictMercos[key00]) - int(dictBmais[key]) < 0:
                    NegativeAndHighStock(f"Produto do Bmais com estoque MAIOR: Cód: {key00} - [BMAIS] Estoque: {dictBmais[key]} // [MERCOS] Estoque: {dictMercos[key00]}") 
            else:
                NotFound(f"Produto do Bmais NÃO encontrado no Mercos: {key}")

    output_workbook.save('C:\\Users\\Usefr\\Desktop\\EstoqueMercos\\Output\\Arquivo de Venda.xlsx')
except Exception as e:
    print(f"An error occurred: {str(e)}")